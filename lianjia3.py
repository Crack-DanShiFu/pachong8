import os

import numpy as np
import pandas as pd
from lxml import etree
from openpyxl import load_workbook, Workbook
import requests
import matplotlib.pyplot as plt

regions = {'dongcheng': 100, 'xicheng': 100, 'chaoyang': 100, 'haidian': 100, 'fengtai': 100, 'shijingshan': 100,
           'tongzhou': 100, 'changping': 100, 'daxing': 100, 'yizhuangkaifaqu': 48, 'shunyi': 100, 'fangshan': 100,
           'mentougou': 49, 'pinggu': 3, 'huairou': 1, 'miyun': 3, 'yanqing': 1,
           }


def getPage(region, index):
    url = 'https://bj.lianjia.com/ershoufang/' + str(region) + '/pg' + str(index)
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36',
    }
    html = requests.get(url, headers=headers).text
    html = etree.HTML(html)
    main_info = html.xpath('//ul[contains(@class,"sellListContent")]/li')
    result = []
    for m in main_info:
        title = m.xpath('div//div[@class="title"]/a/text()')[0]
        houseInfo = m.xpath('div//div[@class="houseInfo"]/text()')[0]
        totalPrice = ''.join(m.xpath('div//div[@class="totalPrice"]//text()'))
        unitPrice = m.xpath('div//div[@class="unitPrice"]//text()')[0]
        address = '-'.join(m.xpath('div//div[@class="positionInfo"]/a/text()'))
        result.append([title, houseInfo, totalPrice, unitPrice, address])
    return result


def getRegionsInfo(region):
    result = []
    for i in range(regions[region]):
        r = getPage(region, i + 1)
        result.extend(r)
        print(region, r)
    # write_excel(region, result)
    return result


def write_excel(region, result):
    path = 'a.xlsx'
    if not os.path.exists(path):
        wb = Workbook()
        wb.save(path)
    wb = load_workbook(path)
    wb.guess_types = True  # 猜测格式类型
    mySheet = wb.create_sheet(title=region)
    for r in result:
        mySheet.append(r)
    wb.save(path)


# for r in regions:
#     getRegionsInfo(r)

def read_excel():
    path = 'a.xlsx'
    if not os.path.exists(path):
        return
    wb = load_workbook(path)
    wb.guess_types = True  # 猜测格式类型
    return wb


def calculation_avg(wb):
    avg_price = {}
    avg_area = {}
    for reg in regions:
        mySheet = wb[reg]
        price = [int(r.value.replace('单价', '').replace('元/平米', '')) for r in mySheet['D']]
        area = [float(r.value.split('|')[1].replace('平米', '')) for r in mySheet['B']]
        avg_price[reg] = sum(price) / len(price)
        avg_area[reg] = sum(area) / len(area)
    return avg_price, avg_area


price_list, area_list = calculation_avg(read_excel())

df1 = pd.DataFrame(price_list.values(), columns=['avg_price'], index=list(price_list.keys()))
df2 = pd.DataFrame(area_list.values(), columns=['avg_area'], index=list(area_list.keys()))
df1.plot.bar()
df2.plot.bar()
plt.show()
